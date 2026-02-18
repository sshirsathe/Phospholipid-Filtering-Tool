import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment


# =============================================================================
# CONFIG: add new lipid classes here (future PI goes here)
# =============================================================================
LIPID_CONFIG = {
    "PE": {
        "filter_cols": ["Name 1"],
        "filter_terms": ["PE"],

        "is_mz_values": [747.7138, 746.7089],
        "is_rt_range": (10, 14),
        "is_area_threshold": 1_000_000,

        "step12_rt_range": (11, 15),
        "step13_rt_range": (10, 20),

        "use_step13a_d6": True,
        "step13a_rt_range": (11, 15),

        "highlight_plus": True,
    },
    "PC": {
        # PC can be in Name 1 OR Name 2
        "filter_cols": ["Name 1", "Name 2"],
        "filter_terms": ["PC"],

        "is_mz_values": [834.764323, 835.7706],
        "is_rt_range": (35, 42),
        "is_area_threshold": 100_000,

        "step12_rt_range": (35, 42),
        "step13_rt_range": (36, 48),

        "use_step13a_d6": False,
        "step13a_rt_range": None,

        "highlight_plus": True,
    },
    "PI": {
        "filter_cols": ["Name 1"],
        "filter_terms": ["PI"],
        "is_mz_values": None,
        "is_rt_range": None,
        "is_area_threshold": None,
        "step12_rt_range": None,
        "step13_rt_range": None,
        "use_step13a_d6": False,
        "step13a_rt_range": None,
        "highlight_plus": True,
    },
}


def dedupe_peak_area_cols_after_avg_blanks_only(
    df_in: pd.DataFrame,
    group_col: str,
    avg_col: str = "Avg_blanks",
    area_prefix: str = "Area:",
    exclude_blank_is: bool = True
) -> pd.DataFrame:
    if df_in.empty:
        return df_in.copy()

    cols = list(df_in.columns)
    if avg_col not in cols:
        raise KeyError(f"'{avg_col}' not found; required to select Area columns after it.")

    avg_idx = cols.index(avg_col)

    eligible_area_cols = []
    for c in cols[avg_idx + 1:]:
        s = str(c)
        if not s.startswith(area_prefix):
            continue
        if exclude_blank_is:
            s_low = s.lower()
            if "blank" in s_low or "is" in s_low:
                continue
        eligible_area_cols.append(c)

    if not eligible_area_cols:
        return df_in.drop_duplicates(subset=[group_col], keep="first").copy()

    df_work = df_in.copy()
    for c in eligible_area_cols:
        df_work[c] = pd.to_numeric(df_work[c], errors="coerce")

    out_rows = []
    for _, g in df_work.groupby(group_col, dropna=False, sort=False):
        g = g.copy()

        rowwise_max = g[eligible_area_cols].max(axis=1, skipna=True)
        rep_idx = rowwise_max.idxmax() if rowwise_max.notna().any() else g.index[0]
        rep_row = g.loc[rep_idx].copy()

        colwise_max = g[eligible_area_cols].max(axis=0, skipna=True)
        for c in eligible_area_cols:
            rep_row[c] = colwise_max[c]

        out_rows.append(rep_row)

    df_out = pd.DataFrame(out_rows)
    return df_out[cols].reset_index(drop=True)


def process_files(input_csv: str, labels_csv: str, output_xlsx: str, lipid_class: str):
    lipid_class = str(lipid_class).strip().upper()
    if lipid_class not in LIPID_CONFIG:
        raise ValueError(f"Unsupported lipid class '{lipid_class}'. Supported: {list(LIPID_CONFIG.keys())}")

    cfg = LIPID_CONFIG[lipid_class]

    required_for_run = ["filter_cols", "filter_terms"]
    for k in required_for_run:
        if not cfg.get(k):
            raise ValueError(f"Lipid class '{lipid_class}' is missing config '{k}'")

    needs_is = True
    if needs_is:
        for k in ["is_mz_values", "is_rt_range", "is_area_threshold", "step12_rt_range", "step13_rt_range"]:
            if cfg.get(k) is None:
                raise ValueError(
                    f"Lipid class '{lipid_class}' is missing '{k}'. "
                    f"Add it to LIPID_CONFIG when you define {lipid_class} filtering."
                )

    input_path = input_csv
    labels_path = labels_csv
    output_csv = output_xlsx.rsplit(".", 1)[0] + ".csv"

    df = pd.read_csv(input_path)

    # ---------- 1) insert m/z after Molecular Weight ----------
    mw_col = "Molecular Weight"
    mz_col = "m/z"
    if mw_col not in df.columns:
        raise KeyError(f"Missing required column: {mw_col}")

    mw_numeric = pd.to_numeric(df[mw_col], errors="coerce")
    df[mz_col] = mw_numeric - 1.007825032

    cols = list(df.columns)
    cols.remove(mz_col)
    mw_idx = cols.index(mw_col)
    cols.insert(mw_idx + 1, mz_col)
    df = df[cols]

    # ---------- 2) group BLANK then IS columns after "Area (Max.)" ----------
    anchor_col = "Area (Max.)"
    if anchor_col not in df.columns:
        raise KeyError(f"Missing required column: {anchor_col}")

    def load_labels_with_header(path: str) -> pd.DataFrame:
        tmp = pd.read_csv(path, header=None)
        header_rows = tmp.index[tmp[0].astype(str).str.strip().eq("File Name")].tolist()
        if not header_rows:
            raise ValueError("Could not find header row 'File Name' in labels file.")
        header_idx = header_rows[0]
        return pd.read_csv(path, skiprows=header_idx, header=0)

    labels_df = load_labels_with_header(labels_path)
    labels_df.columns = [str(c).strip() for c in labels_df.columns]

    required = {"File Name", "Sample ID"}
    missing = required - set(labels_df.columns)
    if missing:
        raise KeyError(f"Labels file missing required columns: {missing}. Found: {list(labels_df.columns)}")

    blank_file_names = (
        labels_df.loc[
            labels_df["Sample ID"].astype(str).str.contains("blank", case=False, na=False),
            "File Name"
        ]
        .astype(str)
        .str.strip()
        .tolist()
    )
    if not blank_file_names:
        raise ValueError("No blank samples found: no Sample ID contained 'blank' in the labels file.")

    all_cols = list(df.columns)
    candidate_cols = [c for c in all_cols if "Area:" in str(c)]

    blank_cols = sorted({
        col
        for fname in blank_file_names
        for col in candidate_cols
        if fname in str(col)
    })

    is_cols = [c for c in all_cols if "_IS_" in c]

    other_cols = [c for c in all_cols if c not in blank_cols + is_cols]
    anchor_idx = other_cols.index(anchor_col)
    prefix_cols = other_cols[:anchor_idx + 1]
    suffix_cols = other_cols[anchor_idx + 1:]

    reordered = prefix_cols + blank_cols + is_cols + suffix_cols
    df = df[reordered]

    # ---------- 3) Avg_blanks if >1 blank col, placed AFTER the last _IS_ col ----------
    avg_col = "Avg_blanks"
    if len(blank_cols) > 1:
        df[avg_col] = df[blank_cols].apply(pd.to_numeric, errors="coerce").mean(axis=1)

        cols = list(df.columns)
        cols.remove(avg_col)

        insert_after = is_cols[-1] if len(is_cols) > 0 else blank_cols[-1]
        ins_idx = cols.index(insert_after) + 1
        cols.insert(ins_idx, avg_col)
        df = df[cols]

    # =========================================================================
    # Helpers: "lipid present?" and "row-wise which Name col contains lipid?"
    # =========================================================================
    filter_cols = cfg["filter_cols"]
    filter_terms = cfg["filter_terms"]
    name1_col = "Name 1"
    name2_col = "Name 2"

    def matches_any_term(series: pd.Series, terms) -> pd.Series:
        s = series.astype(str)
        m = False
        for t in terms:
            m = m | s.str.contains(t, na=False)
        return m

    def lipid_present_mask(df_: pd.DataFrame) -> pd.Series:
        m = False
        for c in filter_cols:
            m = m | matches_any_term(df_[c], filter_terms)
        return m

    def which_name_col_has_lipid(df_: pd.DataFrame) -> pd.Series:
        """
        Returns "Name 1" if lipid term is in Name 1,
                "Name 2" if lipid term is in Name 2 (and not in Name 1),
                else "Name 1" (fallback).
        """
        n1 = matches_any_term(df_[name1_col], filter_terms)
        if name2_col in df_.columns:
            n2 = matches_any_term(df_[name2_col], filter_terms)
        else:
            n2 = pd.Series(False, index=df_.index)

        out = pd.Series(name1_col, index=df_.index, dtype=object)
        out = out.where(~((~n1) & n2), name2_col)
        return out

    # Adds a per-row "active name column" and "active name value"
    df["_active_name_col"] = which_name_col_has_lipid(df)
    df["_active_name"] = df[name1_col].astype(str)
    if name2_col in df.columns:
        use_name2 = df["_active_name_col"].eq(name2_col)
        df.loc[use_name2, "_active_name"] = df.loc[use_name2, name2_col].astype(str)

    # ---------- 5) class-specific filter ----------
    df = df[lipid_present_mask(df)].copy()

    # Need to recompute active-name columns after filtering
    df["_active_name_col"] = which_name_col_has_lipid(df)
    df["_active_name"] = df[name1_col].astype(str)
    if name2_col in df.columns:
        use_name2 = df["_active_name_col"].eq(name2_col)
        df.loc[use_name2, "_active_name"] = df.loc[use_name2, name2_col].astype(str)

    # ---------- 6) sort by RT [min] ascending ----------
    rt_col = "RT [min]"
    if rt_col not in df.columns:
        raise KeyError(f"Missing required column: {rt_col}")

    df[rt_col] = pd.to_numeric(df[rt_col], errors="coerce")
    df = df.sort_values(by=rt_col, ascending=True, na_position="last")

    # =============================================================================
    # Steps 7-9: IS subset
    # For PC: "IS" must be checked in the same Name col where PC was found (active col)
    # =============================================================================
    def mz_ppm_window(mz, ppm):
        delta = mz * ppm / 1_000_000
        return mz - delta, mz + delta

    ppm = 20
    mz_windows = [mz_ppm_window(mz, ppm) for mz in cfg["is_mz_values"]]

    def in_any_ppm_window(mz, windows):
        if pd.isna(mz):
            return False
        return any(lower <= mz <= upper for lower, upper in windows)

    df[mz_col] = pd.to_numeric(df[mz_col], errors="coerce")

    # IS mask depends on lipid class:
    # - PE: original behavior (Name 1)
    # - PC: check IS in the active name value (same Name column where PC is)
    if lipid_class == "PC":
        is_mask = df["_active_name"].astype(str).str.contains("IS", na=False)
    else:
        is_mask = df[name1_col].astype(str).str.contains("IS", na=False)

    df_is = df[is_mask].copy()
    df_is = df_is[df_is[mz_col].apply(lambda x: in_any_ppm_window(x, mz_windows))].copy()

    df_is[rt_col] = pd.to_numeric(df_is[rt_col], errors="coerce")
    is_rt_min, is_rt_max = cfg["is_rt_range"]
    df_is = df_is[df_is[rt_col].between(is_rt_min, is_rt_max, inclusive="both")].copy()

    area_col = "Area (Max.)"
    if area_col not in df_is.columns:
        raise KeyError(f"Missing required column: {area_col}")

    df_is[area_col] = pd.to_numeric(df_is[area_col], errors="coerce")
    df_is = df_is[df_is[area_col].notna() & (df_is[area_col].abs() >= cfg["is_area_threshold"])].copy()

    # Dedupe/grouping:
    # - PE: by Name 1
    # - PC: by _active_name (so Name2-only PC species group correctly)
    group_col = "_active_name" if lipid_class == "PC" else name1_col

    df_is = dedupe_peak_area_cols_after_avg_blanks_only(
        df_is,
        group_col=group_col,
        avg_col=avg_col,
        area_prefix="Area:",
        exclude_blank_is=True
    )

    # Remove IS rows using same mask logic used to build df_is
    df_no_is = df[~is_mask].copy()
    df_final = pd.concat([df_no_is, df_is], ignore_index=True)

    df_final[rt_col] = pd.to_numeric(df_final[rt_col], errors="coerce")
    df_final = df_final.sort_values(by=rt_col, ascending=True, na_position="last")

    # Recompute active-name columns on df_final (needed for steps 11-13)
    df_final["_active_name_col"] = which_name_col_has_lipid(df_final)
    df_final["_active_name"] = df_final[name1_col].astype(str)
    if name2_col in df_final.columns:
        use_name2 = df_final["_active_name_col"].eq(name2_col)
        df_final.loc[use_name2, "_active_name"] = df_final.loc[use_name2, name2_col].astype(str)

    # =============================================================================
    # 11) Clean name: apply semicolon/no-plus cleaning to the ACTIVE name column.
    # PE -> Name 1, PC -> active Name col (Name1 or Name2 per row)
    # =============================================================================
    def clean_name_semicolon_no_plus(series: pd.Series) -> pd.Series:
        s = series.astype(str)
        has_semi = s.str.contains(";", na=False)
        pref = s.str.split(";", n=1).str[0]
        m = has_semi & (~pref.str.contains(r"\+", na=False))
        return s.where(~m, pref)

    if lipid_class == "PC":
        # Clean Name 1 for rows where active col is Name 1
        n1_rows = df_final["_active_name_col"].eq(name1_col)
        df_final.loc[n1_rows, name1_col] = clean_name_semicolon_no_plus(df_final.loc[n1_rows, name1_col])

        # Clean Name 2 for rows where active col is Name 2
        if name2_col in df_final.columns:
            n2_rows = df_final["_active_name_col"].eq(name2_col)
            df_final.loc[n2_rows, name2_col] = clean_name_semicolon_no_plus(df_final.loc[n2_rows, name2_col])
    else:
        df_final[name1_col] = clean_name_semicolon_no_plus(df_final[name1_col])

    # Recompute active name after cleaning
    df_final["_active_name_col"] = which_name_col_has_lipid(df_final)
    df_final["_active_name"] = df_final[name1_col].astype(str)
    if name2_col in df_final.columns:
        use_name2 = df_final["_active_name_col"].eq(name2_col)
        df_final.loc[use_name2, "_active_name"] = df_final.loc[use_name2, name2_col].astype(str)

    # =============================================================================
    # 12) Replace lipid rows (no IS, no '+'): apply '+' check on the ACTIVE name only
    # =============================================================================
    step12_rt_min, step12_rt_max = cfg["step12_rt_range"]

    lip_mask = lipid_present_mask(df_final)

    if lipid_class == "PC":
        is_mask_final = df_final["_active_name"].astype(str).str.contains("IS", na=False)
        plus_mask_final = df_final["_active_name"].astype(str).str.contains(r"\+", na=False)
    else:
        is_mask_final = df_final[name1_col].astype(str).str.contains("IS", na=False)
        plus_mask_final = df_final[name1_col].astype(str).str.contains(r"\+", na=False)

    replace_mask = lip_mask & (~is_mask_final) & (~plus_mask_final)

    df_lip = df_final.loc[replace_mask].copy()
    df_lip[rt_col] = pd.to_numeric(df_lip[rt_col], errors="coerce")
    df_lip = df_lip[df_lip[rt_col].between(step12_rt_min, step12_rt_max, inclusive="both")].copy()

    df_lip = dedupe_peak_area_cols_after_avg_blanks_only(
        df_lip,
        group_col=("_active_name" if lipid_class == "PC" else name1_col),
        avg_col=avg_col,
        area_prefix="Area:",
        exclude_blank_is=True
    )

    df_no_lip = df_final.loc[~replace_mask].copy()
    df_final = pd.concat([df_no_lip, df_lip], ignore_index=True)

    df_final[rt_col] = pd.to_numeric(df_final[rt_col], errors="coerce")
    df_final = df_final.sort_values(by=rt_col, ascending=True, na_position="last")

    # Recompute active name columns after concatenation
    df_final["_active_name_col"] = which_name_col_has_lipid(df_final)
    df_final["_active_name"] = df_final[name1_col].astype(str)
    if name2_col in df_final.columns:
        use_name2 = df_final["_active_name_col"].eq(name2_col)
        df_final.loc[use_name2, "_active_name"] = df_final.loc[use_name2, name2_col].astype(str)

    # =============================================================================
    # 13) For lipid rows with "+": drop those OUTSIDE RT window
    # Use '+' check ONLY on ACTIVE name
    # =============================================================================
    step13_rt_min, step13_rt_max = cfg["step13_rt_range"]

    lip_mask = lipid_present_mask(df_final)
    plus_in_active = df_final["_active_name"].astype(str).str.contains(r"\+", na=False) if lipid_class == "PC" else \
                     df_final[name1_col].astype(str).str.contains(r"\+", na=False)

    plus_mask = lip_mask & plus_in_active
    plus_outside_rt_mask = plus_mask & ~df_final[rt_col].between(step13_rt_min, step13_rt_max, inclusive="both")
    df_final = df_final.loc[~plus_outside_rt_mask].copy()

    # Convert nums to numeric
    for col in df_final.columns:
        if col.startswith("Area:") or col in ["Area (Max.)", "Avg_blanks", "m/z", "RT [min]"]:
            df_final[col] = pd.to_numeric(df_final[col], errors="coerce")

    # =============================================================================
    # 13a) Optional D6 logic: D6 check only on ACTIVE name
    # =============================================================================
    if cfg.get("use_step13a_d6", False):
        d6_min, d6_max = cfg["step13a_rt_range"]
        lip_mask = lipid_present_mask(df_final)

        if lipid_class == "PC":
            d6_mask = df_final["_active_name"].astype(str).str.contains("D6", na=False)
        else:
            d6_mask = df_final[name1_col].astype(str).str.contains("D6", na=False)

        d6_mask = lip_mask & d6_mask
        d6_outside = d6_mask & ~df_final[rt_col].between(d6_min, d6_max, inclusive="both")
        df_final = df_final.loc[~d6_outside].copy()

        for col in df_final.columns:
            if col.startswith("Area:") or col in ["Area (Max.)", "Avg_blanks", "m/z", "RT [min]"]:
                df_final[col] = pd.to_numeric(df_final[col], errors="coerce")

    # =============================================================================
    # 14) Create "Area - Avg_blanks" columns ONLY for Area columns that appear AFTER Avg_blanks
    # =============================================================================
    if avg_col not in df_final.columns:
        raise KeyError(
            f"Step 14 requires '{avg_col}' to exist. "
            f"It is only created when there are >1 blank columns."
        )

    cols = list(df_final.columns)
    avg_idx = cols.index(avg_col)

    eligible_area_cols = []
    for c in cols[avg_idx + 1:]:
        s = str(c)
        if not s.startswith("Area:"):
            continue
        s_low = s.lower()
        if "blank" in s_low or "is" in s_low:
            continue
        eligible_area_cols.append(c)

    if not eligible_area_cols:
        raise ValueError(
            "Step 14: No eligible 'Area:' columns found after Avg_blanks "
            "(excluding any with 'blank' or 'IS' in the name)."
        )

    df_final[avg_col] = pd.to_numeric(df_final[avg_col], errors="coerce")
    for c in eligible_area_cols:
        df_final[c] = pd.to_numeric(df_final[c], errors="coerce")

    new_cols = []
    for c in eligible_area_cols:
        new_name = f"{c} - Avg blanks"
        diff = df_final[c] - df_final[avg_col]
        df_final[new_name] = diff.clip(lower=0)
        new_cols.append(new_name)

    cols = list(df_final.columns)
    last_eligible_idx = max(cols.index(c) for c in eligible_area_cols)

    spacer_col = ""
    if spacer_col in cols:
        spacer_col = " "
    df_final[spacer_col] = ""

    cols = list(df_final.columns)
    cols.remove(spacer_col)
    for c in new_cols:
        cols.remove(c)

    insert_pos = last_eligible_idx + 1
    cols[insert_pos:insert_pos] = [spacer_col] + new_cols
    df_final = df_final[cols]

    # =============================================================================
    # Sort rows A->Z by Name 1 (case-insensitive)
    # =============================================================================
    df_final[name1_col] = df_final[name1_col].astype(str)
    df_final = df_final.sort_values(
        by=name1_col,
        key=lambda s: s.str.upper(),
        na_position="last"
    ).reset_index(drop=True)

    # =============================================================================
    # Save outputs
    # =============================================================================
    df_final.to_csv(output_csv, index=False, float_format="%.10f")
    df_final.to_excel(output_xlsx, index=False)

    # =============================================================================
    # Styling in Excel: highlight rows where ACTIVE name contains '+' AND lipid is present
    # =============================================================================
    lip_mask = lipid_present_mask(df_final)
    plus_in_active = df_final["_active_name"].astype(str).str.contains(r"\+", na=False) if lipid_class == "PC" else \
                     df_final[name1_col].astype(str).str.contains(r"\+", na=False)

    plus_remaining_mask = lip_mask & plus_in_active

    if lipid_class == "PC":
        is_name_mask = df_final["_active_name"].astype(str).str.contains("IS", na=False)
    else:
        is_name_mask = df_final[name1_col].astype(str).str.contains("IS", na=False)

    wb = load_workbook(output_xlsx)
    ws = wb.active

    ws.row_dimensions[1].height = 121.50
    header_align = Alignment(wrap_text=True, vertical="top")
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col_idx).alignment = header_align

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    note_yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    for i in range(len(df_final)):
        excel_row = i + 2
        if bool(is_name_mask.iloc[i]):
            fill = note_yellow_fill
        elif bool(plus_remaining_mask.iloc[i]) and cfg.get("highlight_plus", True):
            fill = red_fill
        else:
            continue

        for c in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=c).fill = fill

    wb.save(output_xlsx)

    print(f"Saved styled Excel: {output_xlsx}")
    print(f"Saved CSV:          {output_csv}")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Lipid Filter Pipeline")
        self.resizable(False, False)

        self.input_path = tk.StringVar()
        self.labels_path = tk.StringVar()
        self.output_xlsx_path = tk.StringVar()
        self.lipid_class = tk.StringVar(value="PE")

        self._build()

    def _build(self):
        pad = {"padx": 10, "pady": 6}

        tk.Label(self, text="Main input CSV").grid(row=0, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.input_path, width=60).grid(row=0, column=1, **pad)
        tk.Button(self, text="Browse", command=self.browse_input).grid(row=0, column=2, **pad)

        tk.Label(self, text="Labels CSV").grid(row=1, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.labels_path, width=60).grid(row=1, column=1, **pad)
        tk.Button(self, text="Browse", command=self.browse_labels).grid(row=1, column=2, **pad)

        tk.Label(self, text="Lipid class").grid(row=2, column=0, sticky="w", **pad)
        tk.OptionMenu(self, self.lipid_class, *LIPID_CONFIG.keys()).grid(row=2, column=1, sticky="w", **pad)

        tk.Label(self, text="Output Excel (.xlsx)").grid(row=3, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.output_xlsx_path, width=60).grid(row=3, column=1, **pad)
        tk.Button(self, text="Choose", command=self.browse_output_xlsx).grid(row=3, column=2, **pad)

        tk.Button(self, text="Run", command=self.run_pipeline, height=2, width=15).grid(row=4, column=1, **pad)

    def browse_input(self):
        p = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if p:
            self.input_path.set(p)

    def browse_labels(self):
        p = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if p:
            self.labels_path.set(p)

    def browse_output_xlsx(self):
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if p:
            self.output_xlsx_path.set(p)

    def run_pipeline(self):
        inp = self.input_path.get().strip()
        lab = self.labels_path.get().strip()
        out_xlsx = self.output_xlsx_path.get().strip()
        lipid = self.lipid_class.get().strip().upper()

        if not inp or not lab or not out_xlsx:
            messagebox.showerror(
                "Missing file",
                "Please select input CSV, labels CSV, and an output Excel (.xlsx) file."
            )
            return

        out_csv = out_xlsx.rsplit(".", 1)[0] + ".csv"

        try:
            process_files(inp, lab, out_xlsx, lipid)
            messagebox.showinfo(
                "Done",
                f"Saved outputs:\n\nExcel (styled):\n{out_xlsx}\n\nCSV (values only):\n{out_csv}"
            )
        except Exception as e:
            messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    App().mainloop()

